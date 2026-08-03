"""Cancellation-aware tabular serialization with a durable, atomic commit.

The writer never mutates an existing destination until a fully serialized and
revalidated temporary file exists.  Cancellation, validation, replacement, and
durability failures all leave the previous bytes in place and produce no
:class:`~universal_rpa.domain.results.OutputCommit`.
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
import shutil
import uuid
from collections.abc import Callable, Sequence
from contextlib import suppress
from importlib import import_module
from io import BytesIO
from pathlib import Path
from typing import Any, Literal, cast
from uuid import UUID

from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from pydantic import BaseModel, ConfigDict, model_validator

from universal_rpa.domain.errors import ErrorCode, RpaError
from universal_rpa.domain.results import LoopCursor, OutputCommit, TableData
from universal_rpa.domain.workflow import OutputRelativePath
from universal_rpa.ports.automation import CancellationToken

TEMP_SUFFIX = ".universal-rpa.tmp"
ROLLBACK_SUFFIX = ".universal-rpa.rollback"

DestinationFlush = Callable[[Path], None]


def canonical_header_hash(headers: Sequence[str]) -> str:
    """Return a stable digest of the exact ordered header row."""

    payload = json.dumps(list(headers), ensure_ascii=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _fsync_path(path: Path) -> None:
    descriptor = os.open(path, os.O_RDWR | getattr(os, "O_BINARY", 0))
    try:
        os.fsync(descriptor)
    finally:
        os.close(descriptor)


def flush_file_write_through(path: Path) -> None:
    """Force *path* to stable storage using Win32 write-through when available."""

    try:
        win32file = cast(Any, import_module("win32file"))
    except ImportError:
        _fsync_path(path)
        return
    try:
        handle = win32file.CreateFile(
            str(path),
            win32file.GENERIC_WRITE,
            win32file.FILE_SHARE_READ,
            None,
            win32file.OPEN_EXISTING,
            win32file.FILE_FLAG_WRITE_THROUGH,
            None,
        )
    except Exception as error:  # pragma: no cover - depends on live filesystem state
        raise OSError("destination cannot be opened for write-through flush") from error
    try:
        win32file.FlushFileBuffers(handle)
    except Exception as error:  # pragma: no cover - depends on live filesystem state
        raise OSError("destination buffers cannot be flushed") from error
    finally:
        handle.Close()


class TableOutputSpec(BaseModel):
    """Where and how one extracted table is committed beneath the output root."""

    model_config = ConfigDict(extra="forbid", frozen=True)

    format: Literal["csv", "xlsx"]
    relative_path: OutputRelativePath
    required_headers: frozenset[str] = frozenset()
    sheet_name: str | None = None

    @model_validator(mode="after")
    def sheet_matches_output_format(self) -> TableOutputSpec:
        if self.format == "csv" and self.sheet_name is not None:
            raise ValueError("CSV output cannot have a sheet name")
        if self.format == "xlsx" and (self.sheet_name is None or not self.sheet_name.strip()):
            raise ValueError("XLSX output requires a nonblank sheet name")
        return self


class AtomicTableWriter:
    def __init__(self, *, flush_destination: DestinationFlush | None = None) -> None:
        self._flush_destination = flush_destination or flush_file_write_through

    def save(
        self,
        table: TableData,
        spec: TableOutputSpec,
        output_root: Path,
        cancellation: CancellationToken,
        producer_step_id: UUID,
        producer_cursor: tuple[LoopCursor, ...],
    ) -> OutputCommit:
        cancellation.raise_if_cancelled()
        self._require_headers(table, spec)
        destination = self._resolve_destination(spec, output_root)
        destination_existed = destination.exists()

        temporary = self._serialize(table, spec, destination, cancellation)
        rollback: Path | None = None
        try:
            cancellation.raise_if_cancelled()
            self._verify(table, spec, temporary)
            if destination_existed:
                rollback = self._copy_rollback(destination)
            cancellation.raise_if_cancelled()
            os.replace(temporary, destination)
        except BaseException:
            self._discard(temporary)
            self._discard(rollback)
            raise

        try:
            self._flush_destination(destination)
            payload = destination.read_bytes()
        except Exception:
            self._restore(destination, rollback, existed=destination_existed)
            raise RpaError(
                ErrorCode.OUTPUT_UNAVAILABLE,
                "출력 파일을 저장 장치에 확정하지 못했습니다.",
            ) from None
        finally:
            self._discard(rollback)

        return OutputCommit(
            destination=destination,
            format=spec.format,
            sheet_name=spec.sheet_name,
            row_count=len(table.rows),
            sha256=hashlib.sha256(payload).hexdigest(),
            headers_sha256=canonical_header_hash(table.headers),
            committed=True,
            producer_step_id=producer_step_id,
            producer_cursor=producer_cursor,
        )

    @staticmethod
    def _require_headers(table: TableData, spec: TableOutputSpec) -> None:
        missing = spec.required_headers - set(table.headers)
        if missing:
            raise RpaError(
                ErrorCode.INVALID_SCHEMA,
                "추출한 표에 필요한 열이 없습니다.",
            )

    @staticmethod
    def _resolve_destination(spec: TableOutputSpec, output_root: Path) -> Path:
        try:
            relative = OutputRelativePath(spec.relative_path.root)
        except Exception:
            raise RpaError(
                ErrorCode.INVALID_SCHEMA,
                "출력 경로가 출력 폴더의 안전한 상대 경로가 아닙니다.",
            ) from None
        root = Path(output_root)
        try:
            root.mkdir(parents=True, exist_ok=True)
        except OSError:
            raise RpaError(ErrorCode.OUTPUT_UNAVAILABLE, "출력 폴더를 만들 수 없습니다.") from None
        try:
            destination = relative.resolve_under(root)
        except (OSError, ValueError):
            raise RpaError(
                ErrorCode.INVALID_SCHEMA,
                "출력 경로가 출력 폴더의 안전한 상대 경로가 아닙니다.",
            ) from None
        try:
            destination.parent.mkdir(parents=True, exist_ok=True)
        except OSError:
            raise RpaError(
                ErrorCode.OUTPUT_UNAVAILABLE, "출력 하위 폴더를 만들 수 없습니다."
            ) from None
        return destination

    def _serialize(
        self,
        table: TableData,
        spec: TableOutputSpec,
        destination: Path,
        cancellation: CancellationToken,
    ) -> Path:
        temporary = destination.parent / f"{destination.name}.{uuid.uuid4().hex}{TEMP_SUFFIX}"
        try:
            if spec.format == "csv":
                self._write_csv(table, temporary, cancellation)
            else:
                self._write_xlsx(table, spec, temporary, cancellation)
            _fsync_path(temporary)
        except BaseException:
            self._discard(temporary)
            raise
        return temporary

    @staticmethod
    def _write_csv(table: TableData, temporary: Path, cancellation: CancellationToken) -> None:
        try:
            with temporary.open("w", encoding="utf-8-sig", newline="") as stream:
                writer = csv.writer(stream, lineterminator="\r\n")
                writer.writerow(table.headers)
                for row in table.rows:
                    cancellation.raise_if_cancelled()
                    writer.writerow(["" if cell is None else cell for cell in row])
                stream.flush()
                os.fsync(stream.fileno())
        except OSError:
            raise RpaError(
                ErrorCode.OUTPUT_UNAVAILABLE, "출력 파일을 기록할 수 없습니다."
            ) from None

    @staticmethod
    def _write_xlsx(
        table: TableData,
        spec: TableOutputSpec,
        temporary: Path,
        cancellation: CancellationToken,
    ) -> None:
        workbook = Workbook(write_only=False)
        try:
            sheet = workbook.active
            sheet.title = spec.sheet_name
            sheet.append(list(table.headers))
            for row in table.rows:
                cancellation.raise_if_cancelled()
                sheet.append(list(row))
            try:
                workbook.save(temporary)
            except OSError:
                raise RpaError(
                    ErrorCode.OUTPUT_UNAVAILABLE, "출력 파일을 기록할 수 없습니다."
                ) from None
        finally:
            workbook.close()

    def _verify(self, table: TableData, spec: TableOutputSpec, temporary: Path) -> None:
        headers, row_count = (
            self._read_csv_shape(temporary)
            if spec.format == "csv"
            else self._read_xlsx_shape(temporary, spec)
        )
        if headers != tuple(table.headers) or row_count != len(table.rows):
            raise RpaError(
                ErrorCode.OUTPUT_UNAVAILABLE,
                "기록한 출력 파일이 추출 결과와 일치하지 않습니다.",
            )

    @staticmethod
    def _read_csv_shape(temporary: Path) -> tuple[tuple[str, ...], int]:
        try:
            with temporary.open("r", encoding="utf-8-sig", newline="") as stream:
                rows = list(csv.reader(stream))
        except (OSError, csv.Error, UnicodeDecodeError):
            raise RpaError(
                ErrorCode.OUTPUT_UNAVAILABLE, "기록한 출력 파일을 다시 읽을 수 없습니다."
            ) from None
        if not rows:
            return (), 0
        return tuple(rows[0]), len(rows) - 1

    @staticmethod
    def _read_xlsx_shape(temporary: Path, spec: TableOutputSpec) -> tuple[tuple[str, ...], int]:
        # openpyxl refuses unknown filename extensions, so the durable temporary
        # file is reopened through a binary stream instead of by path.
        try:
            workbook = load_workbook(BytesIO(temporary.read_bytes()), read_only=True)
        except Exception:
            raise RpaError(
                ErrorCode.OUTPUT_UNAVAILABLE, "기록한 출력 파일을 다시 읽을 수 없습니다."
            ) from None
        try:
            sheet = workbook[spec.sheet_name]
            values = [row for row in sheet.values]
        except Exception:
            raise RpaError(
                ErrorCode.OUTPUT_UNAVAILABLE, "기록한 출력 시트를 다시 읽을 수 없습니다."
            ) from None
        finally:
            workbook.close()
        if not values:
            return (), 0
        return tuple(str(cell) for cell in values[0]), len(values) - 1

    @staticmethod
    def _copy_rollback(destination: Path) -> Path:
        rollback = destination.parent / f"{destination.name}.{uuid.uuid4().hex}{ROLLBACK_SUFFIX}"
        try:
            shutil.copyfile(destination, rollback)
            _fsync_path(rollback)
        except OSError:
            with suppress(OSError):
                rollback.unlink(missing_ok=True)
            raise RpaError(
                ErrorCode.OUTPUT_UNAVAILABLE,
                "기존 출력 파일의 복구본을 만들 수 없습니다.",
            ) from None
        return rollback

    def _restore(self, destination: Path, rollback: Path | None, *, existed: bool) -> None:
        if existed and rollback is not None and rollback.exists():
            with suppress(OSError):
                os.replace(rollback, destination)
            with suppress(Exception):
                self._flush_destination(destination)
            return
        if not existed:
            with suppress(OSError):
                destination.unlink(missing_ok=True)

    @staticmethod
    def _discard(path: Path | None) -> None:
        if path is None:
            return
        with suppress(OSError):
            path.unlink(missing_ok=True)


__all__ = [
    "ROLLBACK_SUFFIX",
    "TEMP_SUFFIX",
    "AtomicTableWriter",
    "DestinationFlush",
    "TableOutputSpec",
    "canonical_header_hash",
    "flush_file_write_through",
]
