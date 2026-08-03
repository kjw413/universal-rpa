from __future__ import annotations

import csv
from collections.abc import Iterator, Sequence
from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]

from universal_rpa.domain.errors import ErrorCode, RpaError
from universal_rpa.domain.types import DataCell, FrozenMapping
from universal_rpa.domain.workflow import (
    CsvDataSource,
    DataSourceSpec,
    InlineDataSource,
    XlsxDataSource,
)
from universal_rpa.ports.data_sources import DataPreview


class TabularDataSourceProvider:
    def preview(
        self,
        project_dir: Path,
        spec: DataSourceSpec,
        max_rows: int = 20,
    ) -> DataPreview:
        if max_rows <= 0:
            raise ValueError("max_rows must be positive")
        headers, rows = self._load(project_dir, spec)
        return DataPreview(
            headers=headers,
            rows=tuple(self._freeze_row(headers, row) for row in rows[:max_rows]),
            total_row_count=len(rows),
        )

    def iter_rows(
        self,
        project_dir: Path,
        spec: DataSourceSpec,
        required_columns: frozenset[str],
    ) -> Iterator[FrozenMapping[str, DataCell]]:
        headers, rows = self._load(project_dir, spec)
        missing = required_columns - frozenset(headers)
        if missing:
            raise RpaError(
                ErrorCode.DATA_SOURCE_INVALID,
                "데이터 소스에 실행에 필요한 열이 없습니다.",
            )
        for row in rows:
            yield self._freeze_row(headers, row)

    def _load(
        self,
        project_dir: Path,
        spec: DataSourceSpec,
    ) -> tuple[tuple[str, ...], tuple[tuple[DataCell, ...], ...]]:
        try:
            if isinstance(spec, InlineDataSource):
                return self._validate_table(spec.headers, spec.rows)
            path = spec.path.resolve_under(project_dir)
            if not path.is_file() or self._is_link_like(path):
                raise ValueError("unsafe input path")
            if isinstance(spec, CsvDataSource):
                return self._read_csv(path, spec.encoding)
            return self._read_xlsx(path, spec)
        except RpaError:
            raise
        except Exception:
            raise RpaError(
                ErrorCode.DATA_SOURCE_INVALID,
                "데이터 파일 형식, 인코딩, 시트 또는 경로를 확인하세요.",
            ) from None

    def _read_csv(
        self,
        path: Path,
        encoding: str,
    ) -> tuple[tuple[str, ...], tuple[tuple[DataCell, ...], ...]]:
        with path.open("r", encoding=encoding, newline="", errors="strict") as stream:
            parsed = tuple(tuple(cell for cell in row) for row in csv.reader(stream))
        if not parsed:
            raise ValueError("empty CSV")
        return self._validate_table(parsed[0], parsed[1:])

    def _read_xlsx(
        self,
        path: Path,
        spec: XlsxDataSource,
    ) -> tuple[tuple[str, ...], tuple[tuple[DataCell, ...], ...]]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if spec.sheet_name not in workbook.sheetnames:
                raise ValueError("missing sheet")
            sheet = workbook[spec.sheet_name]
            parsed = tuple(tuple(self._cell(value) for value in row) for row in sheet.values)
        finally:
            workbook.close()
        if not parsed:
            raise ValueError("empty XLSX")
        return self._validate_table(parsed[0], parsed[1:])

    @staticmethod
    def _cell(value: object) -> DataCell:
        if value is None or isinstance(value, (bool, int, float, str)):
            return value
        raise ValueError("nested or unsupported cell")

    @classmethod
    def _validate_table(
        cls,
        raw_headers: Sequence[object],
        raw_rows: Sequence[Sequence[object]],
    ) -> tuple[tuple[str, ...], tuple[tuple[DataCell, ...], ...]]:
        headers = tuple(str(header).strip() if header is not None else "" for header in raw_headers)
        if (
            not headers
            or any(not header for header in headers)
            or len(set(headers)) != len(headers)
        ):
            raise ValueError("headers must be nonblank and unique")
        rows: list[tuple[DataCell, ...]] = []
        for raw_row in raw_rows:
            if len(raw_row) != len(headers):
                raise ValueError("row width drift")
            rows.append(tuple(cls._cell(value) for value in raw_row))
        return headers, tuple(rows)

    @staticmethod
    def _freeze_row(
        headers: tuple[str, ...],
        row: tuple[DataCell, ...],
    ) -> FrozenMapping[str, DataCell]:
        return FrozenMapping(tuple(zip(headers, row, strict=True)))

    @staticmethod
    def _is_link_like(path: Path) -> bool:
        if path.is_symlink():
            return True
        is_junction = getattr(path, "is_junction", None)
        return bool(is_junction is not None and is_junction())


__all__ = ["TabularDataSourceProvider"]
