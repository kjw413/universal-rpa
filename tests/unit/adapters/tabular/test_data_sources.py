from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from universal_rpa.adapters.tabular import TabularDataSourceProvider
from universal_rpa.domain.errors import RpaError
from universal_rpa.domain.workflow import CsvDataSource, ProjectRelativePath, XlsxDataSource


def test_csv_preview_uses_explicit_encoding_without_guessing(tmp_path: Path) -> None:
    inputs = tmp_path / "inputs"
    inputs.mkdir()
    path = inputs / "rows.csv"
    path.write_bytes("공장\nA\n".encode("cp949"))
    spec = CsvDataSource(
        data_source_id="rows",
        label="행",
        path=ProjectRelativePath("inputs/rows.csv"),
        encoding="utf-8",
    )

    with pytest.raises(RpaError):
        TabularDataSourceProvider().preview(tmp_path, spec)


def test_csv_rows_are_defensively_frozen_and_required_columns_checked(tmp_path: Path) -> None:
    inputs = tmp_path / "inputs"
    inputs.mkdir()
    (inputs / "rows.csv").write_text("factory,date\nA,2026-07-27\n", encoding="utf-8")
    spec = CsvDataSource(
        data_source_id="rows",
        label="행",
        path=ProjectRelativePath("inputs/rows.csv"),
        encoding="utf-8",
    )
    provider = TabularDataSourceProvider()

    preview = provider.preview(tmp_path, spec)
    rows = tuple(provider.iter_rows(tmp_path, spec, frozenset({"factory"})))

    assert preview.headers == ("factory", "date")
    assert dict(rows[0]) == {"factory": "A", "date": "2026-07-27"}
    with pytest.raises(TypeError):
        rows[0]["factory"] = "mutated"  # type: ignore[index]
    with pytest.raises(RpaError):
        tuple(provider.iter_rows(tmp_path, spec, frozenset({"missing"})))


def test_xlsx_requires_existing_named_sheet(tmp_path: Path) -> None:
    inputs = tmp_path / "inputs"
    inputs.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(("factory",))
    sheet.append(("A",))
    workbook.save(inputs / "rows.xlsx")
    spec = XlsxDataSource(
        data_source_id="rows",
        label="행",
        path=ProjectRelativePath("inputs/rows.xlsx"),
        sheet_name="Missing",
    )

    with pytest.raises(RpaError):
        TabularDataSourceProvider().preview(tmp_path, spec)
